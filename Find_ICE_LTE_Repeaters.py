from math import cos, asin, sqrt, pi
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import argparse

def distance(lat1, lon1, lat2, lon2):
    p = pi/180
    a = 0.5 - cos((lat2-lat1)*p)/2 + cos(lat1*p) * cos(lat2*p) * (1-cos((lon2-lon1)*p))/2
    return (12742 * asin(sqrt(a))) #2*R*asin...

dateTimeObj = datetime.now()
print("Started: " + str(dateTimeObj))

parser = argparse.ArgumentParser()
parser.add_argument('-f', '--file', help='Enter the file name')
parser.add_argument('-d', '--distance', type=float, help='takes parameter from reports otherwise 1 km')
args = parser.parse_args()

#  Loading file in data_only more to read parameters from formulae
wb1 = load_workbook(filename=args.file, data_only=True)

sheet1 = wb1["LTE_BINNED_DATA"]
max_row_bins = sheet1["H1"].value

sheet2 = wb1["LTE_REPEATERS"]
max_row_rep = sheet2["H1"].value

wb1.close()

# Load file to make changes.
workbook = load_workbook(filename=args.file)

lte_Bins = workbook["LTE_BINNED_DATA"] # getting LTE BINNED DATA
lte_Repeaters = workbook["LTE_REPEATERS"] # getting LTE_REPEATERS
if args.distance:
    dist_threshold = args.distance
else:
    parameters = workbook['Parameters']
    dist_threshold = parameters["B4"].value
result_Repeaters = []
dis = float
for row1 in lte_Bins.iter_rows(min_row=2, max_row=max_row_bins, min_col=1, max_col=5):
       
    for row2 in lte_Repeaters.iter_rows(min_row=2, max_row=max_row_rep, min_col=1, max_col=5):

        dis = distance(row1[2].value, row1[3].value, row2[3].value, row2[4].value)
                                        
        if dis < dist_threshold:
            result_Repeaters.append(row2[1].value)

    repeatersCount = len(result_Repeaters)
    repeaters = ''
    if repeatersCount > 0:
        for i in range(repeatersCount):
            if i == 0:
                repeaters = result_Repeaters[i]
            else:
                repeaters = repeaters + ", " + result_Repeaters[i]
        lte_Bins["E" + str(row1[0].row)] = repeaters
        result_Repeaters.clear()
    elif len(result_Repeaters) == 0:
        next

workbook.save(filename=args.file)

dateTimeObj = datetime.now()
print("Ended:   " + str(dateTimeObj))
